"""
mapear_rappi_skus.py
====================
Lee el Excel de Rappi (ProductosActualizacion-es.xlsx) y rellena la columna
rappi_product_id en la tabla `productos` para cada SKU que coincida.

El Excel usa el formato Colsports_XXXX en la columna "SKU", donde XXXX es
el SKU de nuestra base de datos (ej. Colsports_2337 → sku='2337').
La columna "ID del producto" contiene el ID único del producto en Rappi.

Uso:
    python scripts/mapear_rappi_skus.py [ruta_excel]

Si no se pasa ruta, busca ProductosActualizacion-es.xlsx en la raíz del proyecto.

Requiere:
    pip install openpyxl
"""

import os
import sys
from typing import Optional

_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _root)

from dotenv import load_dotenv
load_dotenv(dotenv_path=os.path.join(_root, ".env"), override=True)

import openpyxl
from sqlalchemy import create_engine, select, update
from sqlalchemy.orm import Session
from models import Producto

# Columnas del Excel (índice base-0)
COL_RAPPI_PRODUCT_ID = 3   # "ID del producto"
COL_SKU_RAPPI        = 5   # "SKU" (formato Colsports_XXXX)
COL_NOMBRE           = 6   # "Nombre del producto"
FIRST_DATA_ROW       = 4   # las filas 1-3 son cabeceras/instrucciones


def extract_local_sku(rappi_sku: str) -> Optional[str]:
    """Convierte 'Colsports_2337' → '2337'. Retorna None si el formato no coincide."""
    if rappi_sku and rappi_sku.startswith("Colsports_"):
        return rappi_sku[len("Colsports_"):]
    return None


def run(excel_path: str) -> None:
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        print("ERROR: DATABASE_URL no está configurado en .env")
        sys.exit(1)

    engine = create_engine(db_url)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    total_rows = ws.max_row

    updated = 0
    skipped_no_match = []
    skipped_no_sku = 0

    with Session(engine) as session:
        for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
            rappi_product_id = row[COL_RAPPI_PRODUCT_ID]
            rappi_sku        = row[COL_SKU_RAPPI]
            nombre           = row[COL_NOMBRE]

            if not rappi_product_id or not rappi_sku:
                skipped_no_sku += 1
                continue

            local_sku = extract_local_sku(str(rappi_sku))
            if not local_sku:
                skipped_no_sku += 1
                continue

            producto = session.execute(
                select(Producto).where(Producto.sku == local_sku)
            ).scalar_one_or_none()

            if not producto:
                skipped_no_match.append((local_sku, nombre))
                continue

            producto.rappi_product_id = str(rappi_product_id)
            updated += 1

        session.commit()

    print(f"\n=== Resultado del mapeo ===")
    print(f"  Productos actualizados : {updated}")
    print(f"  Filas sin SKU válido   : {skipped_no_sku}")
    print(f"  SKUs no encontrados en DB ({len(skipped_no_match)}):")
    for sku, nombre in skipped_no_match:
        print(f"    - {sku}  ({nombre})")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        path = os.path.join(_root, "ProductosActualizacion-es.xlsx")

    if not os.path.exists(path):
        print(f"ERROR: No se encontró el archivo {path}")
        sys.exit(1)

    print(f"Leyendo: {path}")
    run(path)
